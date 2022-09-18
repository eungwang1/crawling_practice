from ast import keyword
import requests
from bs4 import BeautifulSoup
import time
import pyautogui
from docx import Document
from openpyxl import Workbook
import util
from openpyxl.styles import Alignment

header = {
    'Host': 'www.coupang.com',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3',
}
keyword = pyautogui.prompt('검색어를 입력해주세요.')
total_count = int(pyautogui.prompt('상품의 갯수를 입력해주세요.'))

page = 1
count = 0
wb = Workbook()
ws = wb.create_sheet(keyword)
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 60
ws['A1'] = "순위"
ws['B1'] = "브랜드명"
ws['C1'] = "제품명"
ws['D1'] = "가격"
ws['E1'] = "상세페이지링크"
while True:
    if count == total_count:
        coupang = 'COUPANG'
        util.createDirectory(coupang)
        wb.save(f"./{coupang}/{keyword}.xlsx")
        break
    response = requests.get(
        f'https://www.coupang.com/np/search?component=&q={keyword}&channel=user&page={page}', headers=header)
    page = page + 1
    html_page = response.text
    soup_page = BeautifulSoup(html_page, 'html.parser')
    products = soup_page.select(".search-product-link")
    for product in products:
        if count == total_count:
            break
        isAd = product.select_one(".ad-badge-text")
        isBestSeller = product.select_one(".best-seller-search-product-wrap")
        if isAd or isBestSeller:
            continue
        url = product.attrs['href']
        response = requests.get(
            f"https://www.coupang.com{url}", headers=header)
        html_detail = response.text
        soup_detail = BeautifulSoup(html_detail, "html.parser")
        brand_name = soup_detail.select_one(".prod-brand-name").text.strip()
        product_name = soup_detail.select_one(
            ".prod-buy-header__title").text.strip()
        proudct_price = soup_detail.select_one(".total-price").text.strip()
        if brand_name == '':
            index = product_name.index(' ')
            brand_name = product_name[0:index]
        count = count + 1
        row = count+1
        print("=======링크========")
        print(response.url)
        print("========랭킹==========")
        print(count)
        print("========브랜드=========")
        print(brand_name)
        print("========이름==========")
        print(product_name)
        print("========가격==========")
        print(proudct_price)
        ws[f"A{str(row)}"] = count+1
        ws[f"B{str(row)}"] = brand_name
        ws[f"C{str(row)}"] = product_name
        ws[f"D{str(row)}"] = proudct_price
        ws[f"E{str(row)}"] = url
